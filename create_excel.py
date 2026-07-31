from openpyxl import Workbook, load_workbook

# Read existing file
source = load_workbook("students_results.xlsx")
src = source.active

# Create new workbook
wb = Workbook()
ws = wb.active
ws.title = "Students Results"

# Copy all rows
for row in src.iter_rows(values_only=True):
    ws.append(row)

# Save
wb.save("students_results_new.xlsx")

print("✅ Created students_results_new.xlsx")
print("Students:", ws.max_row - 1)